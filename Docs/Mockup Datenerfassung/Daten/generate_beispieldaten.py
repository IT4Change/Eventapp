# -*- coding: utf-8 -*-
"""Erzeugt ~50 fiktive Beispieldatensätze (spirituelle Events, Rhein-Main-Neckar)
gemäß Referenzen/datenstruktur.xlsx und schreibt XLSX + JSON."""
import json
from datetime import date, time, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HINWEIS = ("Alle Angaben ohne Gewähr. Termin, Ort und Anmeldebedingungen bitte direkt "
           "beim Veranstalter prüfen. Dieses Event wurde automatisch aus öffentlichen "
           "Quellen übernommen und redaktionell aufbereitet.")

# (Name, Teaser, Kategorie, Subkategorie, Ort, Location, Anbieter-Slug,
#  Anmeldung, Quelle, Tage-Dauer, Startzeit, Endzeit, Detailbeschreibung)
E = []
def ev(name, teaser, kat, sub, ort, loc, slug, anm, quelle, dauer, st, en, detail):
    E.append(dict(name=name, teaser=teaser, kat=kat, sub=sub, ort=ort, loc=loc,
                  slug=slug, anm=anm, quelle=quelle, dauer=dauer, st=st, en=en, detail=detail))

K_MUSIK = "Musik & Stimme"
K_TANZ = "Tanz & Bewegung"
K_KLANG = "Klang & Zeremonie"
K_MEDI = "Meditation & Achtsamkeit"
K_RETREAT = "Retreats & Festivals"

ev("Offener Singkreis Frankfurt – Lieder der Erde", "Gemeinsames Singen einfacher Herzenslieder – ohne Vorkenntnisse, ohne Noten.", K_MUSIK, "Singkreis", "Frankfurt am Main", "Kulturraum Lichtblick, Bornheim", "lieder-der-erde", "empfohlen", "Community-Kalender Conscious Rhein-Main", 0, "19:00", "21:30", "In gemütlicher Runde singen wir mehrstimmige Mantren und Herzenslieder aus verschiedenen Kulturen. Die Lieder werden Zeile für Zeile angeleitet, sodass alle sofort mitsingen können. Im Anschluss gibt es Tee und Zeit für Begegnung. Bitte bequeme Kleidung und eine Wasserflasche mitbringen.")
ev("Mantra-Konzert mit Anjali & Friends", "Ein Abend voller Kirtan, Call-and-Response und stiller Momente.", K_MUSIK, "Kirtan / Mantra-Konzert", "Heidelberg", "Yogahaus Neckarblick", "anjali-kirtan", "ja", "Newsletter des Anbieters", 0, "19:30", "22:00", "Anjali & Friends nehmen euch mit auf eine Reise durch klassische Sanskrit-Mantren und eigene Kompositionen. Harmonium, Gitarre und Tabla begleiten den gemeinsamen Gesang, der sich zwischen kraftvollen Höhepunkten und tiefer Stille bewegt. Sitzkissen vorhanden, Spendenempfehlung 15–25 Euro.")
ev("Ecstatic Dance Mannheim – Summer Wave", "Frei tanzen ohne Worte, ohne Schuhe, ohne Alkohol – mit Live-DJ.", K_TANZ, "Ecstatic Dance", "Mannheim", "Alte Seilerei, Halle 3", "ecstatic-mannheim", "empfohlen", "Telegram-Gruppe Ecstatic Dance Rhein-Neckar", 0, "20:00", "23:00", "Ein bewusster Tanzraum mit einer dramaturgisch aufgebauten DJ-Reise von sanften Klängen bis zu treibenden Beats und zurück in die Stille. Wir beginnen mit einem gemeinsamen Opening-Circle. Es gelten die klassischen Ecstatic-Dance-Prinzipien: kein Reden auf der Tanzfläche, keine Substanzen, achtsamer Kontakt.")
ev("Kakaozeremonie & Klangreise bei Vollmond", "Zeremonieller Rohkakao, Gongs und Klangschalen unter dem Vollmond.", K_KLANG, "Kakaozeremonie", "Darmstadt", "Seminarraum Herzraum, Martinsviertel", "herzraum-darmstadt", "ja", "Instagram-Post des Anbieters", 0, "19:00", "22:30", "Wir öffnen den Abend mit einer Schale zeremoniellen Rohkakaos, begleitet von Intention-Setting und sanfter Atemführung. Anschließend trägt euch eine neunzigminütige Klangreise mit Gongs, Klangschalen und Monochord in tiefe Entspannung. Bitte Matte, Decke und Kissen mitbringen. Begrenzte Teilnehmerzahl.")
ev("Klangbad im Salzraum", "Tiefenentspannung mit Klangschalen in besonderer Salzgrotten-Atmosphäre.", K_KLANG, "Klangbad / Sound Healing", "Wiesbaden", "Salzgrotte Aurelia", "klangbad-aurelia", "ja", "Webseite des Anbieters", 0, "18:00", "19:30", "Im mit Himalajasalz ausgekleideten Raum hüllen euch obertonreiche Klangschalen, Koshi-Klangspiele und Ocean Drum in einen Klangteppich. Die salzhaltige Luft unterstützt die Tiefenentspannung. Liegen und Decken sind vorhanden, bitte warme Socken mitbringen.")
ev("5Rhythmen-Welle: Ankommen im Körper", "Angeleitete Bewegungsreise durch die fünf Rhythmen nach Gabrielle Roth.", K_TANZ, "5Rhythmen", "Frankfurt am Main", "Tanzraum Westhafen", "fuenfrhythmen-ffm", "empfohlen", "Newsletter Lichtkreis Rhein-Main", 0, "19:00", "21:30", "Die Welle führt durch Flowing, Staccato, Chaos, Lyrical und Stillness – eine Landkarte, um über Bewegung in Kontakt mit sich selbst zu kommen. Keine Schritte, keine Choreografie, nur dein Körper und die Musik. Offen für alle Erfahrungsstufen.")
ev("Morgenmeditation am Rhein", "Stille Meditation und sanftes Atmen bei Sonnenaufgang am Rheinufer.", K_MEDI, "Stille Meditation", "Mainz", "Rheinufer, Treffpunkt Winterhafen", "stille-am-rhein", "nein", "Community-Kalender Conscious Rhein-Main", 0, "06:30", "07:30", "Wir treffen uns bei (fast) jedem Wetter am Winterhafen und sitzen gemeinsam vierzig Minuten in Stille, eingerahmt von einer kurzen Anleitung und einem Abschlusskreis. Bitte Sitzkissen oder Bank-Hocker sowie warme Kleidung mitbringen. Kostenfrei, Spenden willkommen.")
ev("Breathwork Journey – Atme dich frei", "Verbundenes Atmen mit Musik – kraftvolle Reise zu dir selbst.", K_MEDI, "Breathwork", "Heidelberg", "Studio Atemraum, Weststadt", "atemraum-hd", "ja", "Instagram-Post des Anbieters", 0, "18:30", "21:00", "Conscious Connected Breathing: Über einen verbundenen Atemrhythmus, getragen von einer kuratierten Playlist, können tiefe emotionale Schichten in Bewegung kommen. Es folgt eine lange Integrationsphase mit Stille und Sharing. Nicht geeignet bei Schwangerschaft, Epilepsie und Herz-Kreislauf-Erkrankungen.")
ev("Frauen-Singkreis Mondin", "Singen, Tönen und Stille im Kreis von Frauen.", K_MUSIK, "Singkreis", "Bensheim", "Praxis Lebensklang, Altstadt", "mondin-kreis", "ja", "Aushang im Yogastudio", 0, "19:30", "21:30", "Ein geschützter Raum für Frauen, um die eigene Stimme jenseits von Richtig und Falsch zu erkunden. Wir singen Frauenlieder aus aller Welt, tönen gemeinsam und lauschen der Stille danach. Keine Chorerfahrung nötig. Tee und Kerzenschein inklusive.")
ev("Contact Improvisation Jam", "Offener Jam für achtsame Bewegungs- und Kontaktforschung.", K_TANZ, "Contact Improvisation", "Frankfurt am Main", "Bewegungsraum Sachsenhausen", "ci-jam-ffm", "nein", "Telegram-Gruppe CI Rhein-Main", 0, "19:30", "22:00", "Nach einem dreißigminütigen angeleiteten Warm-up öffnet sich der Raum für freie Contact Improvisation. Anfänger:innen sind herzlich willkommen, erfahrene Tänzer:innen unterstützen. Es gilt ein klares Consent-Prinzip: Jedes Nein ist vollständig.")
ev("Gong-Nacht – Schlafen im Klang", "Eine ganze Nacht getragen von Planetengongs – inkl. Übernachtung im Raum.", K_KLANG, "Gong Bath", "Worms", "Seminarhaus Klangfeld", "klangfeld-worms", "ja", "Newsletter des Anbieters", 1, "21:00", "08:00", "Von 21 Uhr bis Mitternacht spielen zwei Gongmeister durchgehend Planetengongs, danach begleiten sanfte Klänge euren Schlaf bis zum Sonnenaufgang. Gemeinsames Frühstück am Morgen. Bitte Isomatte, Schlafsack und Kissen mitbringen.")
ev("Yin Yoga & Klang – Loslassen am Abend", "Lange gehaltene Yin-Positionen, getragen von Klangschalen.", K_MEDI, "Yin Yoga mit Klang", "Speyer", "Yogaraum Domblick", "domblick-yoga", "ja", "Webseite des Anbieters", 0, "19:00", "20:45", "Stille Yin-Haltungen von drei bis fünf Minuten treffen auf live gespielte Klangschalen. Die Kombination lädt Faszien und Nervensystem zum tiefen Loslassen ein. Für alle Level geeignet, Hilfsmittel sind vorhanden.")
ev("Schwitzhütten-Zeremonie im Odenwald", "Traditionelle Inipi-Zeremonie mit Feuer, Steinen und Gesang.", K_KLANG, "Schwitzhütte", "Michelstadt", "Naturplatz Waldlichtung (genaue Anfahrt nach Anmeldung)", "feuerkreis-odenwald", "ja", "Empfehlung aus der Community", 0, "15:00", "22:00", "Wir bauen gemeinsam das Feuer auf, ehren die vier Runden der Schwitzhütte mit Liedern und Gebeten und schließen mit einem gemeinsamen Essen am Feuer. Bitte Handtuch, Wechselkleidung und eine Schüssel fürs Potluck mitbringen. Vorgespräch verpflichtend.")
ev("Soundbath Deluxe – 60 Gongs & Schalen", "Großes Klangbad mit zwei Spielern und über 60 Instrumenten.", K_KLANG, "Klangbad / Sound Healing", "Mannheim", "Kunsthalle Lichtsaal", "soundbath-deluxe", "ja", "Veranstaltungsportal RheinNeckar-Events", 0, "17:00", "19:00", "Ein immersives Klangerlebnis: Gongs, Kristallklangschalen, Chimes, Monochorde und Stimme verweben sich zu einem zweistündigen Klangraum. Liegeplätze mit Matten sind vorbereitet, eigene Decke empfohlen. Einlass ab 16:30 Uhr, danach geschlossener Raum.")
ev("Achtsamkeitswanderung im Taunus", "Schweigend wandern, bewusst wahrnehmen – Waldbaden-Elemente inklusive.", K_MEDI, "Achtsamkeit in der Natur", "Bad Homburg", "Treffpunkt Parkplatz Hirschgarten", "wald-achtsam", "empfohlen", "Community-Kalender Conscious Rhein-Main", 0, "10:00", "13:30", "Drei Stunden achtsames Gehen im Taunuswald mit Übungen aus dem Waldbaden (Shinrin Yoku): Sinnesöffnung, Schweigestrecken, Baummeditation und ein Abschluss-Tee aus der Thermoskanne. Festes Schuhwerk und wetterfeste Kleidung erforderlich.")
ev("Kirtan-Abend im Ashram-Stil", "Traditionelles Bhajan-Singen mit Harmonium und Mridanga.", K_MUSIK, "Kirtan / Mantra-Konzert", "Hanau", "Zentrum Shanti Mandir", "shanti-hanau", "nein", "Aushang im Bioladen", 0, "18:00", "20:00", "Ein klassischer Satsang-Abend: Wir singen traditionelle Bhajans und Kirtans im Wechselgesang, dazwischen kurze Lesungen aus der Bhagavad Gita. Im Anschluss gibt es Prasad (gesegnetes Essen). Eintritt frei, Spenden willkommen.")
ev("Tanz der Elemente – Erde, Wasser, Feuer, Luft", "Geführte Tanzreise durch die vier Elemente.", K_TANZ, "Bewusster Tanz", "Wiesbaden", "Studio Bewegtes Sein", "bewegtes-sein", "ja", "Newsletter des Anbieters", 0, "19:00", "21:30", "Jedes Element bekommt seinen eigenen musikalischen und energetischen Raum: erdige Rhythmen, fließende Klänge, feurige Beats und luftige Weite. Mit kurzen Embodiment-Anleitungen zwischen den Phasen. Barfuß oder in Schläppchen.")
ev("Stille Retreat-Wochenende: Ankommen", "Zwei Tage Schweigen, Sitzen, Gehen – begleitete Einkehr.", K_RETREAT, "Schweige-Retreat", "Heppenheim", "Seminarhaus Bergstraße", "stille-bergstrasse", "ja", "Webseite des Anbieters", 2, "17:00", "14:00", "Von Freitagabend bis Sonntagmittag üben wir edles Schweigen: Sitzmeditation, Gehmeditation, achtsames Essen und tägliche Einzelgespräche mit der Begleitung. Vegetarische Vollverpflegung und Unterbringung im Einzel- oder Doppelzimmer. Für Anfänger:innen geeignet.")
ev("Conscious Festival Rhein-Main", "Drei Tage Workshops, Musik, Zeremonien und Community.", K_RETREAT, "Festival", "Rüsselsheim", "Festivalgelände Mainwiesen", "conscious-rheinmain", "ja", "Veranstaltungsportal RheinNeckar-Events", 2, "15:00", "16:00", "Über 40 Workshops aus Yoga, Tanz, Atemarbeit und Klang, dazu abendliche Konzerte und Ecstatic-Dance-Floors, ein Healing-Areal und eine Kinderwiese. Camping auf dem Gelände möglich. Frühbucher-Tickets bis sechs Wochen vor Beginn.")
ev("Trommelkreis am Lagerfeuer", "Offenes Trommeln für alle – Instrumente teilweise vorhanden.", K_MUSIK, "Trommelkreis", "Groß-Gerau", "Naturfreundehaus Am Wald", "trommelkreis-gg", "nein", "Telegram-Gruppe Trommelkreis Rhein-Main", 0, "19:00", "22:00", "Wir trommeln gemeinsam am Feuer – von einfachen Grooves bis zu tranceartigen Rhythmen. Einige Djemben und Percussion-Instrumente können ausgeliehen werden. Bei Regen weichen wir in den Saal aus. Bitte Sitzunterlage mitbringen.")
ev("Kundalini Yoga & Mantra-Morgen", "Dynamische Kriyas, Atem und Mantren zum Wochenstart.", K_MEDI, "Kundalini Yoga", "Frankfurt am Main", "Studio Goldenes Tor, Nordend", "goldenes-tor", "ja", "Webseite des Anbieters", 0, "07:00", "08:30", "Eine kraftvolle Morgenpraxis: Kundalini-Kriya, Pranayama, Tiefenentspannung und gemeinsames Mantra-Singen. Danach gibt es Yogi-Tee. Auch für Einsteiger:innen geeignet, bitte bequeme helle Kleidung tragen.")
ev("Cacao & Dance – Herzöffnung in Bewegung", "Kakaozeremonie mit anschließendem Ecstatic Dance.", K_KLANG, "Kakaozeremonie mit Tanz", "Heidelberg", "Halle 02 Studio, Bahnstadt", "cacao-dance-hd", "ja", "Instagram-Post des Anbieters", 0, "18:00", "22:00", "Wir beginnen im Sitzkreis mit zeremoniellem Kakao und Intention, dann öffnet die DJane den Tanzraum für eine zweistündige Reise. Abschluss im stillen Kreis mit Klang. Bitte eigene Tasse mitbringen.")
ev("Familien-Singkreis am Sonntag", "Lieder für Groß und Klein – generationenübergreifendes Singen.", K_MUSIK, "Singkreis", "Darmstadt", "Bürgerhaus Bessungen", "familiensingen-da", "empfohlen", "Community-Kalender Conscious Rhein-Main", 0, "11:00", "12:30", "Einfache Kanons, Bewegungslieder und Herzenslieder für Familien mit Kindern von 0 bis 99. Im Anschluss gemeinsames Picknick im Park (jeder bringt etwas mit). Teilnahme auf Spendenbasis.")
ev("Vision Quest Vorbereitungstag", "Infotag zur viertägigen Visionssuche im Herbst.", K_RETREAT, "Visionssuche", "Weinheim", "Naturcamp Odenwaldrand", "visionsweg", "ja", "Newsletter des Anbieters", 0, "10:00", "17:00", "Wer im Herbst zur viertägigen Visionssuche in die Natur gehen möchte, bekommt an diesem Tag alle Grundlagen: Ablauf, Fasten, Sicherheit, innere Vorbereitung. Mit Council-Runden und einer Solozeit im Wald. Der Tag ist Voraussetzung für die Teilnahme an der Quest.")
ev("Klangmeditation in der Krypta", "Meditative Klänge in tausendjährigem Gemäuer.", K_KLANG, "Klangmeditation", "Lorsch", "Krypta der Königshalle (fiktiv)", "klang-krypta", "ja", "Veranstaltungsportal RheinNeckar-Events", 0, "20:00", "21:30", "Die besondere Akustik der Krypta trägt Stimme, Monochord und Klangschalen in eine fast überirdische Dimension. Begrenzte Plätze, warme Kleidung empfohlen – das Gewölbe ist auch im Sommer kühl.")
ev("Ecstatic Dance Frankfurt – Open Floor", "Der Klassiker am Sonntagabend: bewusst tanzen in großer Halle.", K_TANZ, "Ecstatic Dance", "Frankfurt am Main", "Halle Karlswerk, Ostend", "ecstatic-ffm", "empfohlen", "Telegram-Gruppe Ecstatic Dance Rhein-Main", 0, "18:00", "21:00", "Drei Stunden bewusster Tanzraum mit wechselnden DJs, Opening- und Closing-Circle. Barfuß-Floor, alkohol- und substanzfrei, keine Gespräche auf der Tanzfläche. Komm wie du bist und lass dich von der Musik bewegen.")
ev("Heilsames Singen für Trauernde", "Ein sanfter Raum, in dem Trauer klingen darf.", K_MUSIK, "Heilsames Singen", "Mainz", "Hospizverein Begegnungsraum", "klang-und-trost", "ja", "Empfehlung aus der Community", 0, "17:00", "19:00", "Begleitet von einer ausgebildeten Singleiterin und einer Trauerbegleiterin singen wir tröstende Lieder, tönen gemeinsam und halten Stille. Es darf geweint, gelacht und geschwiegen werden. Begrenzte Teilnehmerzahl, Anmeldung erforderlich.")
ev("Holotropes Atmen – Tagesworkshop", "Intensive Atemarbeit nach Stanislav Grof mit Sitter-Prinzip.", K_MEDI, "Holotropes Atmen", "Mannheim", "Praxis Innenwelt, Neckarau", "innenwelt-ma", "ja", "Webseite des Anbieters", 0, "09:30", "18:30", "In Zweierteams (Atmende und Sitter im Wechsel) erkunden wir erweiterte Bewusstseinszustände durch beschleunigte Atmung, evokative Musik und Körperarbeit. Mit Malphase und Integrationskreis. Ärztlicher Fragebogen vorab erforderlich.")
ev("Neumond-Frauenkreis mit Kakao", "Ritualkreis zu Neumond: Kakao, Journaling, Schwesternschaft.", K_KLANG, "Ritualkreis", "Offenbach", "Raum Rotes Zelt", "rotes-zelt-of", "ja", "Instagram-Post des Anbieters", 0, "19:00", "21:30", "Zum Neumond versammeln wir uns im Kreis: zeremonieller Kakao, geführte Innenschau, Journaling-Impulse und ein Ritual zum Setzen neuer Samen. Offen für alle Frauen, keine Vorerfahrung nötig.")
ev("Biodanza – Tanz des Lebens", "Lebensfreude in Bewegung: Biodanza-Abend für alle.", K_TANZ, "Biodanza", "Aschaffenburg", "Studio Lebenstanz", "biodanza-ab", "empfohlen", "Aushang im Bioladen", 0, "19:30", "21:45", "Biodanza verbindet Musik, Bewegung und Begegnung zu Vivencias – Erlebnissen im Hier und Jetzt. Die Übungen werden kurz angeleitet, getanzt wird ohne Worte. Keine Tanzerfahrung nötig, offene Gruppe.")
ev("Obertongesang lernen – Workshop", "Einführung in Khoomei und westlichen Obertongesang.", K_MUSIK, "Stimm-Workshop", "Wiesbaden", "Musikschule Klangwerk", "oberton-wiesbaden", "ja", "Webseite des Anbieters", 0, "14:00", "18:00", "Schritt für Schritt entdecken wir die Obertöne in der eigenen Stimme: Vokalfärbungen, Zungentechniken und das Spiel mit Resonanzräumen. Maximal zwölf Teilnehmende, Aufnahmen für zuhause inklusive.")
ev("Waldbaden & Klang am Melibokus", "Shinrin Yoku mit Abschluss-Klangmeditation auf der Lichtung.", K_MEDI, "Waldbaden", "Zwingenberg", "Treffpunkt Wanderparkplatz Melibokus", "waldklang-bergstrasse", "ja", "Newsletter Lichtkreis Rhein-Main", 0, "15:00", "18:30", "Langsames, geführtes Eintauchen in den Bergsträßer Wald mit Achtsamkeitsübungen, Schweigephasen und Tee-Zeremonie. Zum Abschluss eine halbstündige Klangmeditation mit Handpan auf der Waldlichtung.")
ev("Tantra-Einführungsabend: Berührung & Präsenz", "Achtsamer Schnupperabend in tantrische Körperarbeit.", K_RETREAT, "Tantra-Workshop", "Heidelberg", "Seminarraum Shakti Loft", "shakti-loft", "ja", "Webseite des Anbieters", 0, "19:00", "22:30", "Ein respektvoller Einführungsabend: Übungen zu Präsenz, Atem und achtsamer Berührung in wechselnden Begegnungen – stets bekleidet und mit klarer Consent-Kultur. Einzeln oder als Paar buchbar.")
ev("Klangschalen-Workshop für Anfänger", "Eigene Klangschale spielen lernen – Technik & Wirkung.", K_KLANG, "Klang-Workshop", "Darmstadt", "Praxis Klangzeit", "klangzeit-da", "ja", "Webseite des Anbieters", 0, "10:00", "13:00", "Anschlagtechniken, Reibetechnik, Wasserklangschale: In kleiner Gruppe lernst du Grundlagen des Klangschalenspiels für Selbstanwendung und Familie. Schalen werden gestellt, eigene dürfen mitgebracht werden.")
ev("Sufi-Abend: Drehen & Dhikr", "Kreisende Bewegung, Atem und Herzensgebet aus der Sufi-Tradition.", K_TANZ, "Sufi / Meditativer Tanz", "Frankfurt am Main", "Saal Derwischhaus, Gallus", "sufi-ffm", "empfohlen", "Empfehlung aus der Community", 0, "19:30", "22:00", "Nach einer Einführung in die Drehtechnik üben wir das meditative Kreisen, begleitet von Live-Musik auf Ney und Rahmentrommel. Dazwischen Dhikr-Gesänge im Kreis. Bequeme weite Kleidung empfohlen.")
ev("Retreat: Yoga & Schweigen an der Bergstraße", "Vier Tage Yoga, Stille und vegetarische Küche.", K_RETREAT, "Yoga-Retreat", "Heppenheim", "Seminarhaus Bergstraße", "stille-bergstrasse", "ja", "Newsletter des Anbieters", 3, "16:00", "13:00", "Täglich zwei Yogaeinheiten, Sitzmeditationen, Schweigen bis zum Mittag und freie Zeit für Spaziergänge in den Weinbergen. Vollverpflegung aus regionaler Bioküche. Begrenzt auf 14 Teilnehmende.")
ev("New Moon Soundjourney", "Klangreise zu Neumond mit Handpan, Stimme und Synthesizer.", K_KLANG, "Klangreise", "Mainz", "Studio Hörraum, Neustadt", "hoerraum-mainz", "ja", "Instagram-Post des Anbieters", 0, "20:00", "21:45", "Eine moderne Klangreise: Handpan, Obertongesang und sphärische Synthesizer-Flächen verschmelzen zu einer Reise nach innen. Liegend auf Matten, begrenzte Platzzahl, Ticketbuchung online.")
ev("Achtsames Bogenschießen", "Meditation in Bewegung: intuitives Bogenschießen im Grünen.", K_MEDI, "Meditation in Bewegung", "Oberursel", "Bogenplatz Taunushang", "bogen-achtsam", "ja", "Webseite des Anbieters", 0, "14:00", "17:00", "Beim intuitiven Bogenschießen verbinden sich Atem, Fokus und Loslassen. Nach der Sicherheitseinweisung üben wir in Stille, mit kurzen Reflexionsrunden. Material wird gestellt, keine Vorerfahrung nötig.")
ev("Männerkreis: Feuer & Stille", "Authentischer Austausch, Schwitzen und Schweigen unter Männern.", K_RETREAT, "Männerkreis", "Weinheim", "Naturcamp Odenwaldrand", "maennerfeuer", "ja", "Telegram-Gruppe Männerarbeit Rhein-Neckar", 0, "16:00", "22:00", "Council-Runde am Feuer, Körperarbeit, Schweigezeit im Wald und ein gemeinsames einfaches Abendessen. Ein Raum für echte Begegnung jenseits von Rollen. Offen für alle Männer ab 18.")
ev("Soundhealing-Konzert: Kristallklangschalen", "Konzertante Klangmeditation mit 432-Hz-Kristallschalen.", K_KLANG, "Klangkonzert", "Ludwigshafen", "Kulturzentrum Hafenlicht", "kristallklang-lu", "ja", "Veranstaltungsportal RheinNeckar-Events", 0, "19:00", "20:30", "Ein Abend mit reinen Kristallklangschalen, gestimmt auf 432 Hz, ergänzt durch Stimme und Chimes. Im Sitzen oder Liegen zu genießen; Matten begrenzt, eigene Decke willkommen.")
ev("Tanzende Meditation: Osho Kundalini", "Die klassische Shaking-Meditation in vier Phasen.", K_TANZ, "Meditativer Tanz", "Darmstadt", "Raum Innehalten, Bessungen", "innehalten-da", "nein", "Community-Kalender Conscious Rhein-Neckar", 0, "18:00", "19:15", "Shaking, freies Tanzen, stilles Sitzen, Liegen: Die Osho Kundalini Meditation löst über sanftes Schütteln Anspannung und macht den Kopf frei. Einfache Anleitung, jeden Mittwoch offen für alle.")
ev("Pilgerwanderung: Jakobsweg-Etappe Rhein-Main", "Geführte Tagesetappe mit spirituellen Impulsen.", K_MEDI, "Pilgern", "Mainz", "Start: Dom-Vorplatz", "pilgern-rheinmain", "ja", "Newsletter Lichtkreis Rhein-Main", 0, "08:00", "17:00", "Eine Tagesetappe auf dem regionalen Jakobsweg mit Schweigestrecken, Impulstexten und Austausch in Gehgemeinschaft. Rund 18 Kilometer, Rucksackverpflegung, Einkehr am Ziel.")
ev("Familienaufstellung & Klang", "Systemische Aufstellungsarbeit, gehalten von Klangraum.", K_RETREAT, "Aufstellungstag", "Frankfurt am Main", "Praxis Wurzelwerk, Bockenheim", "wurzelwerk-ffm", "ja", "Webseite des Anbieters", 0, "10:00", "18:00", "Ein Tagesseminar mit drei bis vier eigenen Aufstellungen und Stellvertreter-Erfahrungen für alle. Klangschalen rahmen die Prozesse und unterstützen die Integration. Vorgespräch für eigene Anliegen erforderlich.")
ev("Chanting & Chai – Mantra-Lounge", "Entspannter Mantra-Abend mit Chai-Bar und Jam-Charakter.", K_MUSIK, "Mantra-Lounge", "Offenbach", "Café Klangoase", "klangoase-of", "nein", "Instagram-Post des Anbieters", 0, "19:00", "22:00", "Halb Konzert, halb offener Jam: Musiker:innen der Region begleiten gemeinsames Mantra-Singen, dazu gibt es hausgemachten Chai und Snacks. Instrumente dürfen mitgebracht werden. Eintritt frei – Hutspende.")
ev("Wim-Hof-Methode: Atmung & Eisbad", "Atemtechnik, Kältetraining und Mindset an einem Tag.", K_MEDI, "Atem & Kälte", "Heidelberg", "Flussbad Neckarwiese (Outdoor)", "eisklar-hd", "ja", "Webseite des Anbieters", 0, "10:00", "15:00", "Nach Theorie und geführten Atemrunden wagen wir uns gemeinsam ins Eisbad – sicher angeleitet und freiwillig. Mit warmem Tee, Theorie zu Kälteadaption und Integrationsrunde. Gesundheits-Check vorab verpflichtend.")
ev("Abendliche Teezeremonie & Stille", "Cha Dao: Tee trinken als Meditation.", K_MEDI, "Teezeremonie", "Wiesbaden", "Teehaus Stiller Garten", "stiller-garten", "ja", "Empfehlung aus der Community", 0, "19:00", "21:00", "In Stille gereichter Tee in mehreren Aufgüssen, dazwischen Raum für das, was sich zeigt. Keine Zeremonie-Erfahrung nötig – nur die Bereitschaft, zwei Stunden langsam zu werden. Maximal acht Gäste.")
ev("Herbst-Equinox-Zeremonie", "Feuerritual und Klang zur Tag-und-Nacht-Gleiche.", K_KLANG, "Jahreskreisfest", "Bensheim", "Streuobstwiese Am Hemsberg", "jahreskreis-bergstrasse", "empfohlen", "Newsletter Lichtkreis Rhein-Main", 0, "18:00", "21:30", "Wir feiern das Gleichgewicht von Licht und Dunkel mit einem Feuerritual, Liedern im Kreis, Dankbarkeitsrunde und abschließender Klangmeditation unter freiem Himmel. Bitte wetterfest kleiden und etwas für das gemeinsame Buffet mitbringen.")
ev("Ecstatic Dance Mannheim – Summer Wave Vol. 2", "Bewusster Tanzabend ohne Worte und Substanzen, mit Live-DJ.", K_TANZ, "Ecstatic Dance", "Mannheim", "Alte Seilerei, Halle 3", "ecstatic-mannheim", "empfohlen", "Veranstaltungsportal RheinNeckar-Events", 0, "20:00", "23:00", "Ein bewusster Tanzraum mit DJ-Reise von sanften Klängen zu treibenden Beats und zurück in die Stille. Opening-Circle zu Beginn. Kein Reden auf der Tanzfläche, keine Substanzen, achtsamer Kontakt.")
ev("Singkreis Lieder der Erde Frankfurt", "Offenes gemeinsames Singen von Herzensliedern ohne Vorkenntnisse.", K_MUSIK, "Singkreis", "Frankfurt am Main", "Kulturraum Lichtblick", "lieder-der-erde", "empfohlen", "Instagram-Post des Anbieters", 0, "19:00", "21:30", "Gemeinsames Singen mehrstimmiger Mantren und Herzenslieder aus verschiedenen Kulturen. Alle Lieder werden angeleitet, Vorkenntnisse sind nicht nötig. Ausklang mit Tee. Bequeme Kleidung empfohlen.")
ev("Lichtmeditation zur Wintersonnenwende", "Gemeinsame Meditation zur längsten Nacht des Jahres.", K_MEDI, "Jahreskreis-Meditation", "Darmstadt", "Seminarraum Herzraum, Martinsviertel", "herzraum-darmstadt", "empfohlen", "Newsletter des Anbieters", 0, "19:00", "21:00", "Zur Wintersonnenwende versammeln wir uns im Kerzenschein zu geführter Lichtmeditation, stillem Sitzen und einem Abschlussritual mit Wunschkarten für das neue Jahr. Mit Tee und Gebäck im Anschluss.")
ev("Handpan-Workshop: Erste Klänge", "Spielerischer Einstieg ins Handpan-Spiel in Kleingruppe.", K_MUSIK, "Instrumenten-Workshop", "Speyer", "Atelier Klangraum Süd", "handpan-speyer", "ja", "Webseite des Anbieters", 0, "11:00", "15:00", "Grundschläge, einfache Patterns und freies Spiel: In maximal acht Personen entdeckst du das Handpan. Instrumente werden gestellt. Am Ende jammen wir gemeinsam im Kreis.")
ev("Frühjahrs-Detox-Retreat: Fasten & Yoga", "Fünf Tage Basenfasten, Yoga und Naturzeit im Odenwald.", K_RETREAT, "Fasten-Retreat", "Michelstadt", "Landhaus Quellengrund", "quellengrund", "ja", "Webseite des Anbieters", 4, "15:00", "11:00", "Basenfasten mit frisch gekochten Brühen und Säften, täglich Yoga und Meditation, geführte Wanderungen und Leberwickel-Workshops. Ärztlich begleitetes Konzept, Vorgespräch inklusive. Einzelzimmer-Option verfügbar.")
ev("Open Stage: Spirit Songs & Poetry", "Offene Bühne für Herzenslieder, Gedichte und Klang.", K_MUSIK, "Open Stage", "Frankfurt am Main", "Café Klangbrücke, Bornheim", "klangbruecke-ffm", "nein", "Community-Kalender Conscious Rhein-Main", 0, "19:30", "22:30", "Eine offene Bühne für alles, was von Herzen kommt: Mantren, eigene Songs, Spoken Word und kleine Klangreisen. Auftrittsslots à zehn Minuten, Anmeldung am Abend vor Ort. Eintritt frei, Hut geht rum.")
ev("Achtsame Massage für Paare – Tagesworkshop", "Berührung lernen: einfache Massagegriffe für zuhause.", "Heilsame Angebote", "Massagen", "Mainz", "Praxis Berührungskunst, Altstadt", "beruehrungskunst-mz", "ja", "Webseite des Anbieters", 0, "10:00", "16:00", "In diesem Tagesworkshop lernt ihr als Paar oder Tandem grundlegende Techniken achtsamer Öl-Massage: Rücken, Nacken, Hände und Füße. Mit Einheiten zu Präsenz, Druck und Kommunikation. Massageliegen, Öle und Mittagssnack inklusive.")

# ---- Remapping auf die offiziellen Kategorien aus Referenzen/kategorien.xlsx ------
K_INSP = "Inspiration & Lernen"
K_HEIL = "Heilsame Angebote"
K_SING = "Singen & Musik"
K_MEHR = "Mehrtägige Events"
REMAP = {  # alte Subkategorie -> (Kategorie, Subkategorie)
    "Singkreis": (K_SING, "Singkreis"),
    "Kirtan / Mantra-Konzert": (K_SING, "Konzert"),
    "Heilsames Singen": (K_SING, "Singkreis"),
    "Trommelkreis": (K_SING, "Singkreis"),
    "Mantra-Lounge": (K_SING, "Konzert"),
    "Open Stage": (K_SING, "Konzert"),
    "Klangreise": (K_SING, "Konzert"),
    "Klangkonzert": (K_SING, "Konzert"),
    "Stimm-Workshop": (K_INSP, "Workshops"),
    "Instrumenten-Workshop": (K_INSP, "Workshops"),
    "Klang-Workshop": (K_INSP, "Workshops"),
    "Aufstellungstag": (K_INSP, "Workshops"),
    "Männerkreis": (K_INSP, "Coachings"),
    "Visionssuche": (K_INSP, "Vorträge"),
    "Ecstatic Dance": ("Tanz", "Ecstatic Dance"),
    "5Rhythmen": ("Tanz", "Ecstatic Dance"),
    "Bewusster Tanz": ("Tanz", "Ecstatic Dance"),
    "Biodanza": ("Tanz", "Ecstatic Dance"),
    "Sufi / Meditativer Tanz": ("Tanz", "Ecstatic Dance"),
    "Contact Improvisation": ("Tanz", "Contact Dance"),
    "Kakaozeremonie": (K_HEIL, "Zeremonien"),
    "Kakaozeremonie mit Tanz": (K_HEIL, "Zeremonien"),
    "Klangbad / Sound Healing": (K_HEIL, "Zeremonien"),
    "Gong Bath": (K_HEIL, "Zeremonien"),
    "Klangmeditation": (K_HEIL, "Zeremonien"),
    "Schwitzhütte": (K_HEIL, "Zeremonien"),
    "Ritualkreis": (K_HEIL, "Zeremonien"),
    "Jahreskreisfest": (K_HEIL, "Zeremonien"),
    "Jahreskreis-Meditation": (K_HEIL, "Zeremonien"),
    "Teezeremonie": (K_HEIL, "Zeremonien"),
    "Breathwork": (K_HEIL, "Breathwork"),
    "Holotropes Atmen": (K_HEIL, "Breathwork"),
    "Atem & Kälte": (K_HEIL, "Breathwork"),
    "Yin Yoga mit Klang": (K_HEIL, "Bewegung"),
    "Kundalini Yoga": (K_HEIL, "Bewegung"),
    "Meditativer Tanz": (K_HEIL, "Bewegung"),
    "Meditation in Bewegung": (K_HEIL, "Bewegung"),
    "Stille Meditation": (K_HEIL, "Natur"),
    "Achtsamkeit in der Natur": (K_HEIL, "Natur"),
    "Waldbaden": (K_HEIL, "Natur"),
    "Pilgern": (K_HEIL, "Natur"),
    "Tantra-Workshop": (K_HEIL, "Körperarbeit"),
    "Schweige-Retreat": (K_MEHR, "Retreats"),
    "Yoga-Retreat": (K_MEHR, "Retreats"),
    "Fasten-Retreat": (K_MEHR, "Retreats"),
    "Festival": (K_MEHR, "Festivals"),
}
for e in E:
    e["kat"], e["sub"] = REMAP.get(e["sub"], (e["kat"], e["sub"]))

# ---- Termine & abgeleitete Felder -------------------------------------------------
START = date(2026, 7, 3)
records = []
seq = 0
for i, e in enumerate(E):
    seq += 1
    d_start = START + timedelta(days=(i * 4) % 175)
    d_end = d_start + timedelta(days=e["dauer"])
    eid = f"{d_start.isoformat()}-{seq:04d}"
    geprueft = i % 3 == 0            # ca. ein Drittel geprüft
    hochgeladen = geprueft and i % 6 == 0  # Teil der geprüften hochgeladen
    records.append({
        "Event ID": eid,
        "Name Event": e["name"],
        "Beschreibung (Zusammenfassung)": e["teaser"],
        "Kategorie": e["kat"],
        "Subkategorie": e["sub"],
        "Datum Start": d_start.isoformat(),
        "Datum Ende": d_end.isoformat(),
        "Startzeit": e["st"],
        "Endzeit": e["en"],
        "Ort": e["ort"],
        "Location": e["loc"],
        "Location Google Maps Link": f"https://maps.example/{e['slug']}",
        "Website Anbieter": f"https://www.{e['slug']}.example",
        "Email Anbieter": f"info@{e['slug']}.example",
        "Telefonnummer Anbieter": f"+49 6{(151 + i * 7) % 900:03d} {900000 + i * 1337}",
        "Anmeldung erfoderlich?": e["anm"],
        "Bild / Flyer vom Anbieter": f"images/{eid}.png",
        "Beschreibung detailliert": e["detail"],
        "Quelle": e["quelle"],
        "Eventhinweis vom Aggregator": HINWEIS,
        "Geprüft": geprueft,
        "Hochgeladen": hochgeladen,
    })

print(f"{len(records)} Datensätze erzeugt")

# ---- JSON --------------------------------------------------------------------------
with open("Daten/events_beispieldaten.json", "w", encoding="utf-8") as f:
    json.dump(records, f, ensure_ascii=False, indent=2)

# ---- XLSX --------------------------------------------------------------------------
wb = Workbook()
ws = wb.active
ws.title = "Events"
headers = list(records[0].keys())
ws.append(headers)
for c in ws[1]:
    c.font = Font(name="Arial", bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", start_color="4A6741")
    c.alignment = Alignment(vertical="center")
for r in records:
    ws.append(["ja" if v is True else ("nein" if v is False else v) for v in r.values()])
for row in ws.iter_rows(min_row=2):
    for c in row:
        c.font = Font(name="Arial", size=10)
        c.alignment = Alignment(vertical="top", wrap_text=False)
widths = {"A": 17, "B": 42, "C": 50, "D": 24, "E": 26, "F": 12, "G": 12, "H": 9, "I": 9,
          "J": 20, "K": 34, "L": 34, "M": 34, "N": 30, "O": 20, "P": 16, "Q": 26,
          "R": 60, "S": 34, "T": 60, "U": 10, "V": 13}
for col, w in widths.items():
    ws.column_dimensions[col].width = w
ws.freeze_panes = "C2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(records) + 1}"
wb.save("Daten/events_beispieldaten.xlsx")
print("XLSX + JSON geschrieben")
